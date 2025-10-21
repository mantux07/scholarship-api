#!/usr/bin/env python3
"""
Author: Tim Smith
Note: All Code owned by Tim

Output Modules for Enhanced Scholarship Agent
Includes: Excel, PDF, HTML, and ICS Calendar generation
"""

import csv
from datetime import datetime
from typing import List
import json


class ExcelExporter:
    """Export scholarships to Excel with formulas and formatting"""

    def __init__(self):
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.utils import get_column_letter
            self.openpyxl = openpyxl
            self.Font = Font
            self.PatternFill = PatternFill
            self.Alignment = Alignment
            self.get_column_letter = get_column_letter
            self.available = True
        except ImportError:
            self.available = False
            print("⚠️  openpyxl not installed. Install with: pip3 install openpyxl")

    def export(self, scholarships: List, filepath: str, student_profile: dict = None) -> bool:
        """Export to Excel with formatting"""
        if not self.available:
            return False

        wb = self.openpyxl.Workbook()
        ws = wb.active
        ws.title = "Scholarships"

        # Headers
        headers = [
            'Priority', 'Scholarship Name', 'Award Amount', 'Min $', 'Max $',
            'Deadline', 'Days Until', 'Min GPA', 'Rec GPA',
            'Essay?', 'Words', 'Rec Letters', 'Interview?',
            'Competition', 'Category', 'Renewable?', 'Est Hours',
            'Application URL', 'Notes'
        ]

        # Write headers with formatting
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.Font(bold=True, color="FFFFFF")
            cell.fill = self.PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = self.Alignment(horizontal="center", vertical="center")

        # Write data
        for row_idx, s in enumerate(scholarships, 2):
            ws.cell(row=row_idx, column=1, value=s.priority_score)
            ws.cell(row=row_idx, column=2, value=s.name)
            ws.cell(row=row_idx, column=3, value=s.amount_display)
            ws.cell(row=row_idx, column=4, value=s.amount_min)
            ws.cell(row=row_idx, column=5, value=s.amount_max)
            ws.cell(row=row_idx, column=6, value=s.deadline)
            ws.cell(row=row_idx, column=7, value=s.days_until_deadline if s.days_until_deadline < 999 else "TBD")
            ws.cell(row=row_idx, column=8, value=s.min_gpa if s.min_gpa > 0 else "None")
            ws.cell(row=row_idx, column=9, value=s.recommended_gpa if s.recommended_gpa > 0 else "N/A")
            ws.cell(row=row_idx, column=10, value="Yes" if s.essay_required else "No")
            ws.cell(row=row_idx, column=11, value=s.essay_word_count if s.essay_word_count > 0 else "N/A")
            ws.cell(row=row_idx, column=12, value=s.rec_letters_required)
            ws.cell(row=row_idx, column=13, value="Yes" if s.interview_required else "No")
            ws.cell(row=row_idx, column=14, value=s.competitiveness)
            ws.cell(row=row_idx, column=15, value=s.category)
            ws.cell(row=row_idx, column=16, value="Yes" if s.renewable else "No")
            ws.cell(row=row_idx, column=17, value=s.estimated_hours)
            ws.cell(row=row_idx, column=18, value=s.application_url)
            ws.cell(row=row_idx, column=19, value=s.notes)

            # Color code by priority score
            priority_cell = ws.cell(row=row_idx, column=1)
            if s.priority_score >= 80:
                priority_cell.fill = self.PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
            elif s.priority_score >= 65:
                priority_cell.fill = self.PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")

        # Adjust column widths
        column_widths = [10, 40, 25, 10, 10, 20, 12, 10, 10, 8, 8, 12, 10, 12, 15, 12, 10, 40, 50]
        for idx, width in enumerate(column_widths, 1):
            ws.column_dimensions[self.get_column_letter(idx)].width = width

        # Freeze first row
        ws.freeze_panes = "A2"

        # Add autofilter
        ws.auto_filter.ref = ws.dimensions

        # Create summary sheet with student profile
        summary = wb.create_sheet("Summary", 0)
        summary['A1'] = "Scholarship Research Summary"
        summary['A1'].font = self.Font(bold=True, size=16)

        # Get student profile info
        profile = student_profile or {}
        row = 3

        # Student Profile Section
        summary[f'A{row}'] = "Student Profile:"
        summary[f'A{row}'].font = self.Font(bold=True, size=12)
        row += 1

        summary[f'A{row}'] = "University:"
        summary[f'B{row}'] = profile.get('university', 'Not specified')
        row += 1

        summary[f'A{row}'] = "Major:"
        summary[f'B{row}'] = profile.get('major', 'Not specified')
        row += 1

        summary[f'A{row}'] = "Year:"
        summary[f'B{row}'] = profile.get('year', 'Not specified')
        row += 1

        summary[f'A{row}'] = "GPA:"
        summary[f'B{row}'] = profile.get('gpa', 'Not specified')
        row += 1

        if profile.get('residency', 'Not specified') != 'Not specified':
            summary[f'A{row}'] = "Residency:"
            summary[f'B{row}'] = profile.get('residency')
            row += 1

        if profile.get('heritage', 'Not specified') != 'Not specified':
            summary[f'A{row}'] = "Heritage:"
            summary[f'B{row}'] = profile.get('heritage')
            row += 1

        if profile.get('gender', 'Not specified') != 'Not specified':
            summary[f'A{row}'] = "Gender:"
            summary[f'B{row}'] = profile.get('gender')
            row += 1

        row += 1  # Blank row

        # Scholarship Statistics Section
        summary[f'A{row}'] = "Scholarship Statistics:"
        summary[f'A{row}'].font = self.Font(bold=True, size=12)
        row += 1

        summary[f'A{row}'] = "Total Scholarships:"
        summary[f'B{row}'] = len(scholarships)
        row += 1

        summary[f'A{row}'] = "Total Potential Award (avg):"
        total_avg = sum((s.amount_min + s.amount_max) / 2 for s in scholarships)
        summary[f'B{row}'] = f"${total_avg:,.0f}"
        row += 1

        summary[f'A{row}'] = "Urgent Deadlines (30 days):"
        urgent = sum(1 for s in scholarships if s.days_until_deadline <= 30 and s.days_until_deadline > 0)
        summary[f'B{row}'] = urgent
        row += 2

        summary[f'A{row}'] = "Generated:"
        summary[f'B{row}'] = datetime.now().strftime('%Y-%m-%d %H:%M')

        wb.save(filepath)
        print(f"✓ Excel file created: {filepath}")
        return True


class PDFExporter:
    """Export scholarships to PDF report"""

    def __init__(self):
        try:
            from reportlab.lib.pagesizes import letter, A4
            from reportlab.lib.styles import getSampleStyleSheet
            from reportlab.lib.units import inch
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
            from reportlab.lib import colors
            self.letter = letter
            self.getSampleStyleSheet = getSampleStyleSheet
            self.inch = inch
            self.SimpleDocTemplate = SimpleDocTemplate
            self.Table = Table
            self.TableStyle = TableStyle
            self.Paragraph = Paragraph
            self.Spacer = Spacer
            self.PageBreak = PageBreak
            self.colors = colors
            self.available = True
        except ImportError:
            self.available = False
            print("⚠️  reportlab not installed. Install with: pip3 install reportlab")

    def export(self, scholarships: List, filepath: str, user_gpa: float, student_profile: dict = None) -> bool:
        """Export top scholarships to PDF"""
        if not self.available:
            return False

        # Get student profile info
        profile = student_profile or {}
        university = profile.get('university', 'Purdue University')
        major = profile.get('major', 'Engineering')
        year = profile.get('year', 'Sophomore')
        heritage = profile.get('heritage', 'Not specified')
        gender = profile.get('gender', 'Not specified')
        gpa = profile.get('gpa', user_gpa)
        residency = profile.get('residency', 'Not specified')

        doc = self.SimpleDocTemplate(filepath, pagesize=self.letter)
        elements = []
        styles = self.getSampleStyleSheet()

        # Title - Dynamic based on student profile
        title = self.Paragraph(f"<b>{university} {major} Scholarship Research Report</b>", styles['Title'])
        elements.append(title)
        elements.append(self.Spacer(1, 0.3*self.inch))

        # Summary - with student profile info
        summary_text = f"""
        <b>Research Date:</b> {datetime.now().strftime('%B %d, %Y')}<br/>
        <b>University:</b> {university}<br/>
        <b>Major:</b> {major}<br/>
        <b>Year:</b> {year}<br/>
        <b>GPA:</b> {gpa}<br/>
        """
        if residency != 'Not specified':
            summary_text += f"<b>Residency:</b> {residency}<br/>"
        if heritage != 'Not specified':
            summary_text += f"<b>Heritage:</b> {heritage}<br/>"
        if gender != 'Not specified':
            summary_text += f"<b>Gender:</b> {gender}<br/>"
        summary_text += f"""<b>Total Scholarships Found:</b> {len(scholarships)}<br/>
        <b>Total Potential Award:</b> ${sum((s.amount_min + s.amount_max)/2 for s in scholarships):,.0f}
        """
        elements.append(self.Paragraph(summary_text, styles['Normal']))
        elements.append(self.Spacer(1, 0.3*self.inch))

        # Top 20 scholarships
        elements.append(self.Paragraph("<b>Top 20 Priority Scholarships</b>", styles['Heading1']))
        elements.append(self.Spacer(1, 0.2*self.inch))

        for i, s in enumerate(scholarships[:20], 1):
            scholarship_text = f"""
            <b>{i}. {s.name}</b><br/>
            <b>Award:</b> {s.amount_display} | <b>Priority Score:</b> {s.priority_score}/100<br/>
            <b>Deadline:</b> {s.deadline} ({s.days_until_deadline} days) | <b>GPA:</b> {s.min_gpa}+ (Rec: {s.recommended_gpa})<br/>
            <b>Category:</b> {s.category} | <b>Competition:</b> {s.competitiveness} | <b>Renewable:</b> {'Yes' if s.renewable else 'No'}<br/>
            <b>Requirements:</b> Essay: {'Yes' if s.essay_required else 'No'}, Rec Letters: {s.rec_letters_required}, Interview: {'Yes' if s.interview_required else 'No'}<br/>
            <b>Est. Time:</b> {s.estimated_hours} hours<br/>
            <b>URL:</b> <link href="{s.application_url}">{s.application_url}</link><br/>
            <b>Notes:</b> {s.notes}
            """
            elements.append(self.Paragraph(scholarship_text, styles['Normal']))
            elements.append(self.Spacer(1, 0.2*self.inch))

            if i % 5 == 0 and i < 20:  # Page break every 5 scholarships
                elements.append(self.PageBreak())

        doc.build(elements)
        print(f"✓ PDF report created: {filepath}")
        return True


class HTMLDashboard:
    """Generate HTML dashboard for scholarship viewing"""

    def export(self, scholarships: List, filepath: str, stats: dict, student_profile: dict = None) -> bool:
        """Generate interactive HTML dashboard"""

        # Get student profile info
        profile = student_profile or {}
        university = profile.get('university', 'Purdue University')
        major = profile.get('major', 'Engineering')
        year = profile.get('year', 'Sophomore')
        heritage = profile.get('heritage', 'Not specified')
        gender = profile.get('gender', 'Not specified')
        gpa = profile.get('gpa', 3.5)
        residency = profile.get('residency', 'Not specified')

        html_content = f"""
<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Purdue Engineering Scholarships - Dashboard</title>
    <style>
        * {{ margin: 0; padding: 0; box-sizing: border-box; }}
        body {{
            font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            padding: 0;
            min-height: 100vh;
        }}
        .browser-header {{
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            color: white;
            text-align: center;
            padding: 20px;
            box-shadow: 0 4px 12px rgba(0,0,0,0.2);
            position: sticky;
            top: 0;
            z-index: 1000;
        }}
        .browser-header h2 {{
            font-size: 1.4em;
            margin-bottom: 12px;
            color: white;
        }}
        .nav-links {{
            display: flex;
            justify-content: center;
            gap: 15px;
            flex-wrap: wrap;
            margin-top: 15px;
        }}
        .nav-links a {{
            color: white;
            text-decoration: none;
            font-weight: 600;
            font-size: 14px;
            padding: 10px 24px;
            background: rgba(255,255,255,0.2);
            border-radius: 25px;
            transition: all 0.3s ease;
            border: 2px solid rgba(255,255,255,0.3);
        }}
        .nav-links a:hover {{
            background: rgba(255,255,255,0.35);
            transform: translateY(-2px);
            box-shadow: 0 4px 12px rgba(0,0,0,0.2);
        }}
        .nav-links .portal-btn {{
            background: rgba(255,255,255,0.95);
            color: #764ba2;
            border: 2px solid white;
            font-size: 15px;
        }}
        .nav-links .portal-btn:hover {{
            background: white;
            color: #667eea;
        }}
        .container {{
            max-width: 1400px;
            margin: 20px auto;
            background: white;
            border-radius: 15px;
            box-shadow: 0 20px 60px rgba(0,0,0,0.3);
            padding: 30px;
        }}
        h1 {{
            color: #333;
            text-align: center;
            margin-bottom: 10px;
        }}
        h2 {{
            color: #764ba2;
            margin-top: 30px;
            margin-bottom: 15px;
            border-left: 4px solid #667eea;
            padding-left: 15px;
        }}
        .subtitle {{
            text-align: center;
            color: #666;
            margin-bottom: 30px;
        }}
        .stats-grid {{
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(200px, 1fr));
            gap: 20px;
            margin-bottom: 30px;
        }}
        .stat-card {{
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            color: white;
            padding: 20px;
            border-radius: 10px;
            text-align: center;
        }}
        .stat-card h3 {{
            font-size: 2em;
            margin-bottom: 10px;
        }}
        .stat-card p {{
            font-size: 0.9em;
            opacity: 0.9;
        }}
        .filters {{
            margin: 20px 0;
            padding: 20px;
            background: #f5f5f5;
            border-radius: 8px;
        }}
        .filters h3 {{
            color: #667eea;
            margin-bottom: 15px;
            font-size: 1.2em;
        }}
        .filters input, .filters select {{
            margin: 0 10px;
            padding: 8px;
            border: 1px solid #ddd;
            border-radius: 4px;
        }}
        table {{
            width: 100%;
            border-collapse: collapse;
            margin-top: 20px;
        }}
        th {{
            background: #667eea;
            color: white;
            padding: 12px;
            text-align: left;
            position: sticky;
            top: 0;
        }}
        td {{
            padding: 10px;
            border-bottom: 1px solid #ddd;
        }}
        tr:hover {{
            background: #f5f5f5;
        }}
        .priority-high {{
            background: #90EE90 !important;
        }}
        .priority-medium {{
            background: #FFFF99 !important;
        }}
        .urgent {{
            color: #e74c3c;
            font-weight: bold;
        }}
        .badge {{
            display: inline-block;
            padding: 3px 8px;
            border-radius: 12px;
            font-size: 0.85em;
            font-weight: bold;
        }}
        .badge-low {{ background: #90EE90; color: #1e5631; }}
        .badge-medium {{ background: #FFD700; color: #7a5c00; }}
        .badge-high {{ background: #FFA07A; color: #8b2500; }}
        .badge-very-high {{ background: #FF6347; color: white; }}
        a {{ color: #667eea; text-decoration: none; }}
        a:hover {{ text-decoration: underline; }}
        .footer {{
            text-align: center;
            margin-top: 30px;
            padding-top: 20px;
            border-top: 2px solid #eee;
            color: #666;
        }}
    </style>
</head>
<body>
    <div class="browser-header">
        <h2>🎓 Scholarship Research Dashboard</h2>
        <div class="nav-links">
            <a href="https://www.purdue.edu/financialaid/scholarships/" target="_blank" class="portal-btn">🎓 Open Purdue Scholarship Portal</a>
            <a href="#stats">Statistics</a>
            <a href="#urgent">Urgent Deadlines</a>
            <a href="#table">All Scholarships</a>
            <a href="#filters">Filters</a>
        </div>
    </div>

    <div class="container">
        <h1 id="stats">🎓 {university} {major} Scholarships</h1>
        <p class="subtitle">{year} | GPA: {gpa} | {residency}</p>
        {f'<p class="subtitle" style="margin-top: -10px;">Heritage: {heritage} | Gender: {gender}</p>' if heritage != 'Not specified' or gender != 'Not specified' else ''}

        <div class="stats-grid">
            <div class="stat-card">
                <h3>{stats['total_scholarships']}</h3>
                <p>Total Scholarships</p>
            </div>
            <div class="stat-card">
                <h3>{stats['gpa_eligible']}</h3>
                <p>GPA Eligible</p>
            </div>
            <div class="stat-card">
                <h3>{stats['urgent_deadlines_30_days']}</h3>
                <p>Urgent (30 Days)</p>
            </div>
            <div class="stat-card">
                <h3>{stats['total_potential_award']}</h3>
                <p>Total Potential</p>
            </div>
        </div>

        <h2 id="urgent">🚨 Urgent Deadlines (Next 30 Days)</h2>
        <p style="margin-bottom: 20px; color: #e74c3c; font-weight: bold;">Apply to these scholarships immediately!</p>

        <div class="filters" id="filters">
            <h3>🔍 Filter & Search Tools</h3>
            <label>Search: <input type="text" id="searchBox" placeholder="Search scholarships..." onkeyup="filterTable()"></label>
            <label>Category:
                <select id="categoryFilter" onchange="filterTable()">
                    <option value="">All Categories</option>
                    <option value="Purdue">Purdue</option>
                    <option value="National">National</option>
                    <option value="Corporate">Corporate</option>
                    <option value="Diversity">Diversity</option>
                    <option value="Professional Org">Professional Org</option>
                    <option value="Out-of-State">Out-of-State</option>
                </select>
            </label>
            <label>Competition:
                <select id="compFilter" onchange="filterTable()">
                    <option value="">All Levels</option>
                    <option value="Low">Low</option>
                    <option value="Medium">Medium</option>
                    <option value="High">High</option>
                    <option value="Very High">Very High</option>
                </select>
            </label>
        </div>

        <h2 id="table">📋 All Scholarships</h2>
        <table id="scholarshipTable">
            <thead>
                <tr>
                    <th>Priority</th>
                    <th>Scholarship Name</th>
                    <th>Award</th>
                    <th>Deadline</th>
                    <th>Days</th>
                    <th>GPA</th>
                    <th>Competition</th>
                    <th>Category</th>
                    <th>Link</th>
                </tr>
            </thead>
            <tbody>
"""

        for s in scholarships:
            priority_class = "priority-high" if s.priority_score >= 80 else ("priority-medium" if s.priority_score >= 65 else "")
            urgent_class = "urgent" if s.days_until_deadline <= 30 and s.days_until_deadline > 0 else ""

            comp_badge_class = {
                'Low': 'badge-low',
                'Medium': 'badge-medium',
                'High': 'badge-high',
                'Very High': 'badge-very-high'
            }.get(s.competitiveness, 'badge-medium')

            days_display = s.days_until_deadline if s.days_until_deadline < 999 else "TBD"

            html_content += f"""
                <tr class="{priority_class}">
                    <td><strong>{s.priority_score}</strong></td>
                    <td><strong>{s.name}</strong><br/>
                        <small>Renewable: {'✓' if s.renewable else '✗'} |
                        Essay: {'✓' if s.essay_required else '✗'} |
                        Letters: {s.rec_letters_required} |
                        Est: {s.estimated_hours}h</small>
                    </td>
                    <td>{s.amount_display}</td>
                    <td class="{urgent_class}">{s.deadline}</td>
                    <td class="{urgent_class}">{days_display}</td>
                    <td>{s.min_gpa}+ / {s.recommended_gpa}</td>
                    <td><span class="badge {comp_badge_class}">{s.competitiveness}</span></td>
                    <td>{s.category}</td>
                    <td><a href="{s.application_url}" target="_blank">Apply</a></td>
                </tr>
"""

        html_content += """
            </tbody>
        </table>

        <div class="footer">
            <p>Generated by Enhanced Scholarship Research Agent</p>
            <p>Author: Tim Smith | All Code owned by Tim</p>
            <p>Research Date: """ + datetime.now().strftime('%B %d, %Y') + """</p>
        </div>
    </div>

    <script>
        function filterTable() {
            const searchValue = document.getElementById('searchBox').value.toLowerCase();
            const categoryValue = document.getElementById('categoryFilter').value;
            const compValue = document.getElementById('compFilter').value;
            const table = document.getElementById('scholarshipTable');
            const rows = table.getElementsByTagName('tr');

            for (let i = 1; i < rows.length; i++) {
                const row = rows[i];
                const name = row.cells[1].textContent.toLowerCase();
                const category = row.cells[7].textContent;
                const competition = row.cells[6].textContent;

                let showRow = true;

                if (searchValue && !name.includes(searchValue)) {
                    showRow = false;
                }
                if (categoryValue && category !== categoryValue) {
                    showRow = false;
                }
                if (compValue && !competition.includes(compValue)) {
                    showRow = false;
                }

                row.style.display = showRow ? '' : 'none';
            }
        }
    </script>
</body>
</html>
"""

        with open(filepath, 'w', encoding='utf-8') as f:
            f.write(html_content)

        print(f"✓ HTML dashboard created: {filepath}")
        return True


class CalendarGenerator:
    """Generate ICS calendar file with scholarship deadlines"""

    def export(self, scholarships: List, filepath: str) -> bool:
        """Generate .ics calendar file"""

        ics_content = """BEGIN:VCALENDAR
VERSION:2.0
PRODID:-//Purdue Scholarship Agent//EN
CALSCALE:GREGORIAN
METHOD:PUBLISH
X-WR-CALNAME:Scholarship Deadlines
X-WR-TIMEZONE:America/New_York
X-WR-CALDESC:Deadlines for Purdue Engineering Scholarships
"""

        for s in scholarships:
            if s.deadline_date:
                # Create event for deadline day
                event_date = s.deadline_date.strftime('%Y%m%d')
                uid = f"{s.name.replace(' ', '-')}-{event_date}@purdue-scholarships.local"

                ics_content += f"""
BEGIN:VEVENT
UID:{uid}
DTSTAMP:{datetime.now().strftime('%Y%m%dT%H%M%SZ')}
DTSTART;VALUE=DATE:{event_date}
SUMMARY:DEADLINE: {s.name}
DESCRIPTION:{s.name}\\n\\nAward: {s.amount_display}\\nGPA: {s.min_gpa}+\\n\\nURL: {s.application_url}\\n\\nNotes: {s.notes}
LOCATION:{s.application_url}
STATUS:CONFIRMED
PRIORITY:5
BEGIN:VALARM
TRIGGER:-P7D
DESCRIPTION:Reminder: {s.name} deadline in 7 days
ACTION:DISPLAY
END:VALARM
END:VEVENT
"""

                # Create reminder 30 days before
                if s.days_until_deadline > 30:
                    reminder_date = (s.deadline_date.replace(day=1) - timedelta(days=30)).strftime('%Y%m%d')
                    reminder_uid = f"{s.name.replace(' ', '-')}-reminder-{reminder_date}@purdue-scholarships.local"

                    ics_content += f"""
BEGIN:VEVENT
UID:{reminder_uid}
DTSTAMP:{datetime.now().strftime('%Y%m%dT%H%M%SZ')}
DTSTART;VALUE=DATE:{reminder_date}
SUMMARY:REMINDER: {s.name} (30 days)
DESCRIPTION:Reminder: {s.name} deadline in 30 days\\n\\nDeadline: {s.deadline}\\nAward: {s.amount_display}\\n\\nStart preparing application materials.
STATUS:CONFIRMED
PRIORITY:3
END:VEVENT
"""

        ics_content += "END:VCALENDAR\n"

        with open(filepath, 'w', encoding='utf-8') as f:
            f.write(ics_content)

        print(f"✓ Calendar file created: {filepath}")
        print("  Import this file into your calendar app (Google Calendar, Apple Calendar, Outlook)")
        return True


class ApplicationTracker:
    """Application tracking system"""

    def create_tracker(self, scholarships: List, filepath: str) -> bool:
        """Create application tracking CSV"""

        fieldnames = [
            'Scholarship Name', 'Award Amount', 'Deadline', 'Days Until',
            'Priority Score', 'Application Status', 'Date Started',
            'Date Submitted', 'Essay Complete', 'Rec Letters Secured',
            'Transcript Sent', 'Interview Scheduled', 'Result',
            'Amount Awarded', 'Notes'
        ]

        with open(filepath, 'w', newline='', encoding='utf-8') as csvfile:
            writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
            writer.writeheader()

            for s in scholarships:
                writer.writerow({
                    'Scholarship Name': s.name,
                    'Award Amount': s.amount_display,
                    'Deadline': s.deadline,
                    'Days Until': s.days_until_deadline if s.days_until_deadline < 999 else 'TBD',
                    'Priority Score': s.priority_score,
                    'Application Status': 'Not Started',
                    'Date Started': '',
                    'Date Submitted': '',
                    'Essay Complete': 'No' if s.essay_required else 'N/A',
                    'Rec Letters Secured': f'0/{s.rec_letters_required}',
                    'Transcript Sent': 'No',
                    'Interview Scheduled': 'No' if s.interview_required else 'N/A',
                    'Result': 'Pending',
                    'Amount Awarded': '',
                    'Notes': ''
                })

        print(f"✓ Application tracker created: {filepath}")
        print("  Use this file to track your application progress")
        return True


from datetime import timedelta
